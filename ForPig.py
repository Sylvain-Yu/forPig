#!/usr/bin/python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

# 可修改路径及sheet_name
path = r'abc.xlsx'
sheet_name = '???'
# 固定语句不需改动
wb = load_workbook(filename = path, data_only = True)
sheet = wb[sheet_name]

# 计算人员姓名位置
# 上个月的位置参数,必须为姓名,非数字
colx1 = 'A'
start_rowx1 = 56
end_rowx1 = 75
# 这个月的人员的位置参数
colx2 = 'A'
start_rowx2 = 77
end_rowx2 = 96
# 复制对应列
col_from_list = ['J','L','H']
col_to_list = ['P','G','C']
row_from = 101
row_to = 77
# 保存的位置，修改单引号内参数
save_path = r'abc.xlsx'

# ————————以下不需要改动——————————
# 获取数据
last_month_persons = []
this_month_persons = []
for x in sheet[colx1+str(start_rowx1):colx1+str(end_rowx1)]:
	last_month_persons.append(x[0].value.strip())
for x in sheet[colx2+str(start_rowx2):colx1+str(end_rowx2)]:
	this_month_persons.append(x[0].value.strip())
quit_persons_list = []
new_persons_list = []

# 查询是否有人员变动
for i in last_month_persons:
	if i not in this_month_persons:
		print(i + '请确认是否离职!')
		quit_persons_list.append(i)
for i in this_month_persons:
	if i not in last_month_persons:
		print(i + '请确认是否为新同事!')
		new_persons_list = []

gap = end_rowx2 - start_rowx2
wb = load_workbook(filename = path, data_only = False)
sheet = wb[sheet_name]
# 复制数据的第一格位置
# 确认列
if len(col_from_list) != len(col_to_list):
	print('请确认复制与粘贴位置相等！')
else:
	for i in range(len(col_from_list)):
		col_from_x = col_from_list[i]
		col_to_x = col_to_list[i]
		basic = sheet[col_from_x + str(row_from):col_from_x + str(row_from + gap)]
		towhere = sheet[col_to_x + str(row_to):col_to_x + str(row_to +gap)]
		# 数据复制 具体操作
		for j in range(gap+1):
			towhere[j][0].value = basic[j][0].value

# 保存文档
wb.save(save_path)
